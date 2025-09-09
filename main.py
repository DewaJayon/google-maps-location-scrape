from playwright.sync_api import sync_playwright
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
import time

def scrape_maps(url):
    data = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()

        page.set_extra_http_headers({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Accept-Language": "en-US,en;q=0.9",
        })

        page.goto(url, wait_until="domcontentloaded")
        time.sleep(3)

        page.wait_for_selector('div[role="feed"]', timeout=20000)

        scroll_container = page.query_selector('div[role="feed"]')

        prev_height = 0
        same_count = 0

        while True:
            page.evaluate(
                """el => el.scrollBy(0, el.scrollHeight)""",
                scroll_container
            )
            time.sleep(2)

            new_height = page.evaluate("el => el.scrollHeight", scroll_container)
            if new_height == prev_height:
                same_count += 1
            else:
                same_count = 0

            if same_count >= 3: 
                break

            prev_height = new_height

        html = page.content()
        soup = BeautifulSoup(html, "html.parser")

        contents = soup.find_all("div", {"class": "Nv2PK"})

        for content in contents:
            map_url = content.find("a", {"class": "hfpxzc"})["href"]

            name = content.find("div", {"class": "qBF1Pd"}).text.strip()

            info_tags = content.find_all("div", {"class": "W4Efsd"})
            content_type, address = None, None

            for info in info_tags:
                text = info.get_text(" ", strip=True)
                if "·" in text:
                    parts = [t.strip() for t in text.split("·")]
                    if len(parts) == 2:
                        content_type, address = parts
                    break

            data.append({
                "Map URL": map_url,
                "Name": name,
                "Type": content_type,
                "Address": address,
            })

    return data


if __name__ == "__main__":
    
    try:
        url = input("Masukkan URL: ")
    except:
        print("Terjadi kesalahan, silahkan coba lagi.")

    try:
        file_name = input("Masukkan nama file untuk menyimpan data: ")
    except:
        print("Mohon masukkan nama file.")
    
    scrape = scrape_maps(url)

    df = pd.DataFrame(scrape)
    df.to_excel(f"{file_name}.xlsx", index=False, engine="openpyxl")
    
    # wb = load_workbook(file_name)
    # ws = wb.active

    # last_row = ws.max_row + 2 
    # ws.cell(row=last_row, column=1, value=f"Total Data: {len(scrape)}")
    # ws.cell(row=last_row+1, column=1, value="Sumber data: Google Maps")
    # ws.cell(row=last_row+2, column=1, value="Disusun oleh: Dewa Jayon")
    # ws.cell(row=last_row+3, column=1, value="Ini hanya data yang terdapat di google maps, metode scraping data, untuk lebih lengkap bisa ditanya di kelurahan penatih.")

    # wb.save(file_name)